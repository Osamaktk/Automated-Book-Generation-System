import json
import csv
import io
import re

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from supabase import Client

from backend.ai import call_ai, stream_ai_async
from backend.config import logger, supabase
from backend.database import (
    create_book,
    create_chapter,
    create_outline,
    get_approved_chapters,
    get_approved_outline,
    get_book,
    get_latest_outline,
    get_pending_review_chapter,
    list_books,
    update_book_status,
    update_outline,
)
from backend.models import BookInput, EditorFeedback
from backend.prompts import (
    build_chapter_prompt,
    build_outline_prompt,
    resolve_planned_chapter_count,
    extract_chapter_title,
)
from backend.services.compiler import compile_to_docx, compile_to_pdf
from backend.services.notifications import notify


router = APIRouter(prefix="/books", tags=["books"])


def _serialize_event(payload: dict) -> str:
    return f"data: {json.dumps(payload)}\n\n"


def _book_notes(book: dict) -> str:
    return book.get("notes_on_outline_before") or book.get("notes") or ""


def _map_outline_status(book_status: str, outline_status: str | None) -> str:
    if outline_status == "approved" or book_status in {"outline_approved", "chapters_in_progress", "chapters_complete"}:
        return "approved"
    if outline_status == "needs_revision":
        return "needs_revision"
    if outline_status == "waiting_for_review" or book_status == "waiting_for_review":
        return "waiting_for_review"
    return outline_status or "not_started"


def _with_project_brief_fields(book: dict, outline: dict | None) -> dict:
    brief_fields = {
        "notes_on_outline_before": _book_notes(book),
        "status_outline_notes": book.get("status_outline_notes")
        or _map_outline_status(book.get("status", ""), outline.get("status") if outline else None),
        "no_notes_needed": book.get("no_notes_needed")
        if book.get("no_notes_needed") is not None
        else book.get("status") == "chapters_complete",
    }
    return {**book, **brief_fields}


def _build_previous_summaries(approved_chapters: list[dict]) -> list[dict]:
    return [
        {"chapter_number": chapter["chapter_number"], "summary": chapter["summary"]}
        for chapter in approved_chapters
        if chapter.get("summary")
    ]


def _sync_book_completion(book: dict, client: Client) -> dict:
    outline = get_latest_outline(book["id"], client=client)
    if not outline:
        return book

    planned_chapter_count = resolve_planned_chapter_count(
        _book_notes(book),
        outline.get("content", ""),
    )
    approved_chapters = get_approved_chapters(book["id"], client=client)
    if (
        planned_chapter_count > 0
        and len(approved_chapters) >= planned_chapter_count
        and book.get("status") != "chapters_complete"
    ):
        update_book_status(book["id"], "chapters_complete", client=client)
        return {**book, "status": "chapters_complete"}
    return book


def _with_planned_chapter_count(book: dict, outline: dict | None) -> dict:
    if not outline:
        return _with_project_brief_fields({**book, "planned_chapter_count": 0}, None)
    planned_chapter_count = resolve_planned_chapter_count(
        _book_notes(book),
        outline.get("content", ""),
        outline.get("editor_notes", ""),
    )
    return _with_project_brief_fields(
        {**book, "planned_chapter_count": planned_chapter_count},
        outline,
    )


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _extract_row_payload(row: dict, index: int) -> dict:
    title = str(row.get("title", "")).strip()
    notes = str(row.get("notes", "")).strip()
    if not title:
        raise HTTPException(
            400,
            f"Spreadsheet row {index} is missing a title. Expected columns: title, notes",
        )
    return {"title": title, "notes": notes}


def _parse_csv_rows(content: bytes) -> list[dict]:
    decoded = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))
    if not reader.fieldnames:
        raise HTTPException(400, "CSV file must include a header row")

    reader.fieldnames = [_normalize_header(name) for name in reader.fieldnames]
    rows = []
    for raw_row in reader:
        normalized = {
            _normalize_header(str(key)): (value or "")
            for key, value in raw_row.items()
            if key is not None
        }
        if any(str(value).strip() for value in normalized.values()):
            rows.append(normalized)
    return rows


def _parse_xlsx_rows(content: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        raise HTTPException(400, "Spreadsheet is empty")

    headers = [_normalize_header(str(cell or "")) for cell in values[0]]
    if not any(headers):
        raise HTTPException(400, "Spreadsheet must include a header row")

    rows = []
    for row in values[1:]:
        normalized = {
            headers[index]: "" if cell is None else str(cell)
            for index, cell in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(str(value).strip() for value in normalized.values()):
            rows.append(normalized)
    return rows


def _load_import_rows(filename: str, content: bytes) -> list[dict]:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "csv":
        return _parse_csv_rows(content)
    if extension == "xlsx":
        return _parse_xlsx_rows(content)
    raise HTTPException(400, "Only .csv and .xlsx files are supported")


def _generate_next_chapter_for_book(book: dict, client: Client) -> dict:
    """Generate the next chapter for a visible book using the caller's RLS client."""
    if book["status"] not in {"outline_approved", "chapters_in_progress"}:
        raise HTTPException(
            400, f"Outline must be approved first. Status: {book['status']}"
        )

    outline = get_approved_outline(book["id"], client=client)
    if not outline:
        raise HTTPException(404, "No approved outline found")

    pending_chapter = get_pending_review_chapter(book["id"], client=client)
    if pending_chapter:
        raise HTTPException(
            400,
            f"Chapter {pending_chapter['chapter_number']} is waiting for review!",
        )

    approved_chapters = get_approved_chapters(book["id"], client=client)
    planned_chapter_count = resolve_planned_chapter_count(
        _book_notes(book),
        outline.get("content", ""),
    )
    if planned_chapter_count > 0 and len(approved_chapters) >= planned_chapter_count:
        update_book_status(book["id"], "chapters_complete", client=client)
        raise HTTPException(
            400,
            "All planned chapters are already complete for this book.",
        )

    next_number = len(approved_chapters) + 1
    chapter_title = extract_chapter_title(outline["content"], next_number)
    previous_summaries = _build_previous_summaries(approved_chapters)
    content = call_ai(
        build_chapter_prompt(
            title=book["title"],
            outline=outline["content"],
            chapter_number=next_number,
            chapter_title=chapter_title,
            previous_summaries=previous_summaries,
        ),
        3000,
    )
    if not content or len(content.strip()) < 100:
        raise HTTPException(500, "AI returned empty content. Please try generating again.")
    chapter = create_chapter(
        book_id=book["id"],
        chapter_number=next_number,
        content=content,
        status="waiting_for_review",
        client=client,
    )
    notify(
        "chapter_ready",
        book["title"],
        f"Chapter {next_number} is ready for review",
    )
    update_book_status(book["id"], "chapters_in_progress", client=client)
    return {
        "message": f"Chapter {next_number} generated!",
        "chapter_id": chapter["id"],
        "chapter_number": next_number,
        "content": content,
        "status": "waiting_for_review",
    }


@router.post("/import")
async def import_books_route(
    file: UploadFile = File(...),
    generate_outlines: bool = Query(
        default=True,
        description="Generate outlines immediately for imported books",
    ),
):
    try:
        filename = file.filename or ""
        content = await file.read()
        if not filename:
            raise HTTPException(400, "Uploaded file must have a filename")
        if not content:
            raise HTTPException(400, "Uploaded file is empty")

        raw_rows = _load_import_rows(filename, content)
        if not raw_rows:
            raise HTTPException(400, "No import rows found in the uploaded file")

        imported_books = []
        for index, row in enumerate(raw_rows, start=2):
            payload = _extract_row_payload(row, index)
            book = create_book(
                payload["title"],
                payload["notes"],
                status="generating" if generate_outlines else "waiting_for_review",
                client=supabase,
            )

            outline_id = None
            outline_status = None
            if generate_outlines:
                outline_content = call_ai(
                    build_outline_prompt(payload["title"], payload["notes"]),
                    2000,
                )
                outline = create_outline(
                    book_id=book["id"],
                    content=outline_content,
                    status="waiting_for_review",
                    client=supabase,
                )
                outline_id = outline["id"]
                outline_status = outline["status"]
                update_book_status(book["id"], "waiting_for_review", client=supabase)
                notify("outline_ready", book["title"], f"Book ID: {book['id']}")

            imported_books.append(
                {
                    "book_id": book["id"],
                    "title": book["title"],
                    "status": "waiting_for_review" if generate_outlines else book["status"],
                    "outline_id": outline_id,
                    "outline_status": outline_status,
                }
            )

        return {
            "message": f"Imported {len(imported_books)} book(s) from spreadsheet",
            "generate_outlines": generate_outlines,
            "books": imported_books,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Book import failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))


@router.post("/create-stream")
async def create_book_stream(
    input: BookInput,
):
    async def generate():
        try:
            book = create_book(
                input.title,
                input.notes,
                status="generating",
                client=supabase,
            )
            book_id = book["id"]

            yield _serialize_event({"type": "book_id", "book_id": book_id})

            prompt = build_outline_prompt(input.title, input.notes)
            full_content = ""

            async for chunk in stream_ai_async(prompt, max_tokens=2000):
                full_content += chunk
                yield _serialize_event({"type": "chunk", "text": chunk})

            outline = create_outline(
                book_id=book_id,
                content=full_content,
                status="waiting_for_review",
                client=supabase,
            )
            notify("outline_ready", book["title"], f"Book ID: {book_id}")
            update_book_status(book_id, "waiting_for_review", client=supabase)

            yield _serialize_event(
                {"type": "done", "book_id": book_id, "outline_id": outline["id"]}
            )
        except Exception as exc:
            logger.error("Outline stream error: %s", exc, exc_info=True)
            yield _serialize_event({"type": "error", "message": str(exc)})

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.post("/create")
async def create_book_route(
    input: BookInput,
):
    try:
        book = create_book(
            input.title,
            input.notes,
            status="generating",
            client=supabase,
        )
        outline_content = call_ai(build_outline_prompt(input.title, input.notes), 2000)
        outline = create_outline(
            book_id=book["id"],
            content=outline_content,
            status="waiting_for_review",
            client=supabase,
        )
        notify("outline_ready", book["title"], f"Book ID: {book['id']}")
        update_book_status(book["id"], "waiting_for_review", client=supabase)
        return {
            "message": "Book created!",
            "book_id": book["id"],
            "outline_id": outline["id"],
            "outline": outline_content,
            "status": "waiting_for_review",
        }
    except Exception as exc:
        logger.error("Create book failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))


@router.get("")
async def list_books_route():
    try:
        books = [
            _sync_book_completion(book, supabase)
            for book in list_books(client=supabase)
        ]
        books = [
            _with_planned_chapter_count(
                book,
                get_latest_outline(book["id"], client=supabase),
            )
            for book in books
        ]
        return {"books": books}
    except Exception as exc:
        logger.error("List books failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))


@router.get("/{book_id}")
async def get_book_route(
    book_id: str,
):
    try:
        book = get_book(book_id, client=supabase)
        if not book:
            raise HTTPException(404, "Book not found")
        book = _sync_book_completion(book, supabase)
        outline = get_latest_outline(book_id, client=supabase)
        book = _with_planned_chapter_count(book, outline)
        return {"book": book, "outline": outline}
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Get book failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))


@router.delete("/{book_id}")
async def delete_book_route(
    book_id: str,
):
    try:
        book = get_book(book_id, client=supabase)
        if not book:
            raise HTTPException(404, "Book not found")

        chapters = (
            supabase.table("chapters").select("id").eq("book_id", book_id).execute().data
            or []
        )
        outlines = (
            supabase.table("outlines").select("id").eq("book_id", book_id).execute().data
            or []
        )

        chapters_deleted = len(chapters)
        outlines_deleted = len(outlines)

        if chapters_deleted:
            supabase.table("chapters").delete().eq("book_id", book_id).execute()

        if outlines_deleted:
            supabase.table("outlines").delete().eq("book_id", book_id).execute()

        supabase.table("books").delete().eq("id", book_id).execute()

        logger.info(
            "Deleted book %s with %s chapters and %s outlines",
            book_id,
            chapters_deleted,
            outlines_deleted,
        )
        return {
            "message": "Book deleted successfully",
            "book_id": book_id,
            "chapters_deleted": chapters_deleted,
            "outlines_deleted": outlines_deleted,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Delete book failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))


@router.post("/{book_id}/feedback")
async def submit_outline_feedback(
    book_id: str,
    feedback: EditorFeedback,
):
    try:
        if feedback.status not in {"approved", "needs_revision"}:
            raise HTTPException(400, "Status must be 'approved' or 'needs_revision'")

        outline = get_latest_outline(book_id, client=supabase)
        if not outline:
            raise HTTPException(404, "Outline not found")

        book = get_book(book_id, client=supabase)
        if not book:
            raise HTTPException(404, "Book not found")

        if feedback.status == "approved":
            update_outline(
                outline["id"],
                client=supabase,
                status="approved",
                editor_notes=feedback.editor_notes,
            )
            update_book_status(book_id, "outline_approved", client=supabase)
            first_chapter = _generate_next_chapter_for_book(
                {**book, "status": "outline_approved"},
                supabase,
            )
            return {
                "message": "Outline approved and first chapter generated!",
                "book_id": book_id,
                "status": "chapters_in_progress",
                "chapter_id": first_chapter["chapter_id"],
                "chapter_number": first_chapter["chapter_number"],
                "next_step": f"Review Chapter {first_chapter['chapter_number']}",
            }

        update_outline(
            outline["id"],
            client=supabase,
            status="needs_revision",
            editor_notes=feedback.editor_notes,
        )
        new_content = call_ai(
            build_outline_prompt(book["title"], _book_notes(book), feedback.editor_notes),
            2000,
        )
        new_outline = create_outline(
            book_id=book_id,
            content=new_content,
            status="waiting_for_review",
            client=supabase,
        )
        notify("outline_ready", book["title"], f"Book ID: {book_id}")
        update_book_status(book_id, "waiting_for_review", client=supabase)
        return {
            "message": "Outline regenerated!",
            "book_id": book_id,
            "outline_id": new_outline["id"],
            "new_outline": new_content,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Outline feedback failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))


@router.get("/{book_id}/compile")
async def compile_book_route(
    book_id: str,
    format: str = "docx",
):
    try:
        if format not in {"docx", "pdf", "txt"}:
            raise HTTPException(400, "Format must be one of: docx, pdf, txt")

        book = get_book(book_id, client=supabase)
        if not book:
            raise HTTPException(404, "Book not found")
        if book["status"] != "chapters_complete":
            raise HTTPException(
                400,
                "Final manuscript export is only available after the book is marked complete.",
            )

        outline = get_latest_outline(book_id, client=supabase)
        if not outline:
            raise HTTPException(404, "Outline not found")

        chapters = (
            supabase.table("chapters").select("*").eq("book_id", book_id).execute().data or []
        )
        approved_chapters = sorted(
            [chapter for chapter in chapters if chapter["status"] == "approved"],
            key=lambda chapter: chapter["chapter_number"],
        )
        if not approved_chapters:
            raise HTTPException(400, "No approved chapters found to compile")

        slug = (
            re.sub(
                r"-+",
                "-",
                re.sub(r"[^a-z0-9]+", "-", book["title"].lower()),
            ).strip("-")
            or "book"
        )
        headers = {
            "Content-Disposition": f'attachment; filename="{slug}_{book_id[:8]}.{format}"'
        }

        if format == "docx":
            content = compile_to_docx(book, outline, approved_chapters)
            logger.info("Compiled DOCX document for book %s", book_id)
            return StreamingResponse(
                iter([content]),
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers=headers,
            )

        if format == "pdf":
            content = compile_to_pdf(book, outline, approved_chapters)
            logger.info("Compiled PDF document for book %s", book_id)
            return StreamingResponse(
                iter([content]),
                media_type="application/pdf",
                headers=headers,
            )

        if format == "txt":
            from backend.services.compiler import compile_to_txt
            content = compile_to_txt(book, outline, approved_chapters)
            logger.info("Compiled TXT document for book %s", book_id)
            return StreamingResponse(
                iter([content]),
                media_type="text/plain; charset=utf-8",
                headers={**headers, "Content-Disposition": f'attachment; filename="{slug}_{book_id[:8]}.txt"'},
            )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Compile book failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))


@router.post("/{book_id}/generate-chapter-stream")
async def generate_chapter_stream(
    book_id: str,
):
    async def generate():
        try:
            book = get_book(book_id, client=supabase)
            if not book:
                yield _serialize_event({"type": "error", "message": "Book not found"})
                return

            if book["status"] not in {"outline_approved", "chapters_in_progress"}:
                yield _serialize_event(
                    {
                        "type": "error",
                        "message": f"Outline must be approved first. Status: {book['status']}",
                    }
                )
                return

            outline = get_approved_outline(book_id, client=supabase)
            if not outline:
                yield _serialize_event(
                    {"type": "error", "message": "No approved outline found"}
                )
                return

            pending_chapter = get_pending_review_chapter(book_id, client=supabase)
            if pending_chapter:
                yield _serialize_event(
                    {
                        "type": "error",
                        "message": f"Chapter {pending_chapter['chapter_number']} is waiting for review!",
                    }
                )
                return

            approved_chapters = get_approved_chapters(book_id, client=supabase)
            planned_chapter_count = resolve_planned_chapter_count(
                _book_notes(book),
                outline.get("content", ""),
            )
            if planned_chapter_count > 0 and len(approved_chapters) >= planned_chapter_count:
                update_book_status(book_id, "chapters_complete", client=supabase)
                yield _serialize_event(
                    {
                        "type": "error",
                        "message": "All planned chapters are already complete for this book.",
                    }
                )
                return
            next_number = len(approved_chapters) + 1
            chapter_title = extract_chapter_title(outline["content"], next_number)
            previous_summaries = _build_previous_summaries(approved_chapters)

            yield _serialize_event(
                {
                    "type": "chapter_info",
                    "chapter_number": next_number,
                    "chapter_title": chapter_title,
                }
            )

            prompt = build_chapter_prompt(
                title=book["title"],
                outline=outline["content"],
                chapter_number=next_number,
                chapter_title=chapter_title,
                previous_summaries=previous_summaries,
            )
            full_content = ""

            async for chunk in stream_ai_async(prompt, max_tokens=3000):
                full_content += chunk
                yield _serialize_event({"type": "chunk", "text": chunk})

            if not full_content or len(full_content.strip()) < 100:
                yield _serialize_event({
                    "type": "error",
                    "message": "AI returned empty or too-short content. Please try generating again."
                })
                return

            chapter = create_chapter(
                book_id=book_id,
                chapter_number=next_number,
                content=full_content,
                status="waiting_for_review",
                client=supabase,
            )
            notify(
                "chapter_ready",
                book["title"],
                f"Chapter {next_number} is ready for review",
            )
            update_book_status(book_id, "chapters_in_progress", client=supabase)

            yield _serialize_event(
                {
                    "type": "done",
                    "chapter_id": chapter["id"],
                    "chapter_number": next_number,
                }
            )
        except Exception as exc:
            logger.error("Chapter stream error: %s", exc, exc_info=True)
            yield _serialize_event({"type": "error", "message": str(exc)})

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.post("/{book_id}/generate-chapter")
async def generate_next_chapter(
    book_id: str,
):
    try:
        book = get_book(book_id, client=supabase)
        if not book:
            raise HTTPException(404, "Book not found")
        return _generate_next_chapter_for_book(book, supabase)
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Generate chapter failed: %s", exc, exc_info=True)
        raise HTTPException(500, str(exc))
